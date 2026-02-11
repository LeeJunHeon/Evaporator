# ui/material_catalog_dialog.py
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional, List, Dict

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QMessageBox, QAbstractItemView, QHeaderView
)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


@dataclass
class MaterialRow:
    material: str
    density_g_cm3: float
    z_factor: float
    note: str = ""


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() == "none":
        return ""
    return s


def _to_float(v: Any, *, default: float = 0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return default
    # "1.000*" 같은 표기 처리
    s = s.replace("*", "").strip()
    # "--------" 같은 표기 처리
    if set(s) <= {"-"}:
        return default
    try:
        return float(s)
    except Exception:
        return default


class MaterialCatalogDialog(QDialog):
    """
    - 최초 실행(저장 파일 없을 때): 물질표.xlsx를 읽어서 config/material_catalog.json 생성
    - 이후 실행: config/material_catalog.json을 읽어서 표시
    - 더블클릭 편집(셀 수정)하면 즉시 JSON 저장(영구 반영)
    - 행 선택 후 Apply 누르면 selected()로 선택 값 반환
    """

    def __init__(self, *, base_dir: Path, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Material Catalog")
        self.setModal(True)
        self.resize(720, 520)

        self._base_dir = Path(base_dir)
        self._config_dir = self._base_dir / "config"
        self._config_dir.mkdir(parents=True, exist_ok=True)

        self._json_path = self._config_dir / "material_catalog.json"
        self._selected: Optional[MaterialRow] = None

        self._loading = True  # itemChanged 저장 루프 방지

        # ---------- UI ----------
        root = QVBoxLayout(self)

        self.infoLabel = QLabel(
            "더블클릭으로 Density/Z-Factor/Material/Note를 수정하면 즉시 저장됩니다.\n"
            "행을 선택한 뒤 Apply를 누르면 Process 탭에 반영됩니다."
        )
        self.infoLabel.setWordWrap(True)
        root.addWidget(self.infoLabel)

        self.table = QTableWidget(self)
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Material", "Density (g/cm³)", "Z factor", "Note"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        root.addWidget(self.table, 1)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)

        self.applyBtn = QPushButton("Apply", self)
        self.cancelBtn = QPushButton("Cancel", self)
        btn_row.addWidget(self.applyBtn)
        btn_row.addWidget(self.cancelBtn)

        root.addLayout(btn_row)

        # ---------- signals ----------
        self.applyBtn.clicked.connect(self._on_apply)
        self.cancelBtn.clicked.connect(self.reject)
        self.table.itemChanged.connect(self._on_item_changed)

        # ---------- load data ----------
        materials, notes = self._load_or_seed()
        if notes:
            # 엑셀 우측 비고(설명)들을 그대로 안내로도 표시
            self.infoLabel.setText(
                self.infoLabel.text()
                + "\n\n[엑셀 비고 요약]\n- "
                + "\n- ".join(notes)
            )

        self._populate(materials)
        self._loading = False

    # ------------------------
    # public
    # ------------------------
    def selected(self) -> Optional[MaterialRow]:
        return self._selected

    # ------------------------
    # internal
    # ------------------------
    def _find_excel_path(self) -> Optional[Path]:
        candidates = [
            self._base_dir / "물질표.xlsx",
            self._base_dir / "config" / "물질표.xlsx",
            self._base_dir / "data" / "물질표.xlsx",
        ]
        for p in candidates:
            if p.exists():
                return p
        return None

    def _load_or_seed(self) -> tuple[list[MaterialRow], list[str]]:
        """
        1) JSON 있으면 JSON 로드
        2) 없으면 Excel(물질표.xlsx) 로드 → JSON 저장(seed)
        """
        if self._json_path.exists():
            try:
                obj = json.loads(self._json_path.read_text(encoding="utf-8"))
                mats = []
                for it in obj.get("materials", []):
                    mats.append(
                        MaterialRow(
                            material=_to_str(it.get("material")),
                            density_g_cm3=float(it.get("density_g_cm3", 0.0)),
                            z_factor=float(it.get("z_factor", 0.0)),
                            note=_to_str(it.get("note")),
                        )
                    )
                notes = [s for s in obj.get("notes", []) if _to_str(s)]
                mats = [m for m in mats if m.material]  # 빈 행 제거
                return mats, notes
            except Exception as e:
                QMessageBox.warning(self, "Load error", f"material_catalog.json 로드 실패:\n{e}")

        # JSON 없으면 Excel seed
        excel_path = self._find_excel_path()
        if excel_path is None:
            QMessageBox.warning(
                self,
                "Missing file",
                "물질표.xlsx를 찾을 수 없습니다.\n"
                "아래 위치 중 하나에 파일을 두세요:\n"
                f"- {self._base_dir / '물질표.xlsx'}\n"
                f"- {self._base_dir / 'config' / '물질표.xlsx'}\n"
                f"- {self._base_dir / 'data' / '물질표.xlsx'}",
            )
            return [], []

        if load_workbook is None:
            QMessageBox.warning(self, "Missing dependency", "openpyxl이 없어 물질표.xlsx를 읽을 수 없습니다.")
            return [], []

        mats, notes = self._read_excel(excel_path)
        # seed 저장
        self._save_json(mats, notes)
        return mats, notes

    def _read_excel(self, excel_path: Path) -> tuple[list[MaterialRow], list[str]]:
        wb = load_workbook(excel_path)
        ws = wb.active

        mats: list[MaterialRow] = []
        notes: list[str] = []

        # 헤더: [물질, Density, Z factor, (빈), Note]
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            material = _to_str(row[0] if len(row) > 0 else "")
            density = _to_float(row[1] if len(row) > 1 else None, default=0.0)
            zfac = _to_float(row[2] if len(row) > 2 else None, default=0.0)
            note = _to_str(row[4] if len(row) > 4 else "")

            if note:
                notes.append(note)

            # 빈 줄 스킵
            if not material:
                continue

            mats.append(MaterialRow(material=material, density_g_cm3=density, z_factor=zfac, note=note))

        # notes 중복 제거(순서 유지)
        uniq = []
        seen = set()
        for n in notes:
            if n not in seen:
                uniq.append(n)
                seen.add(n)

        return mats, uniq

    def _populate(self, mats: list[MaterialRow]) -> None:
        self.table.setRowCount(len(mats))
        for r, m in enumerate(mats):
            it0 = QTableWidgetItem(m.material)
            it1 = QTableWidgetItem(f"{m.density_g_cm3:g}")
            it2 = QTableWidgetItem(f"{m.z_factor:g}")
            it3 = QTableWidgetItem(m.note or "")

            # 정렬/선택 편의
            it0.setData(Qt.UserRole, m.material)

            self.table.setItem(r, 0, it0)
            self.table.setItem(r, 1, it1)
            self.table.setItem(r, 2, it2)
            self.table.setItem(r, 3, it3)

    def _collect_rows(self) -> tuple[list[MaterialRow], list[str]]:
        mats: list[MaterialRow] = []
        notes: list[str] = []

        for r in range(self.table.rowCount()):
            material = _to_str(self.table.item(r, 0).text() if self.table.item(r, 0) else "")
            density = _to_float(self.table.item(r, 1).text() if self.table.item(r, 1) else None, default=0.0)
            zfac = _to_float(self.table.item(r, 2).text() if self.table.item(r, 2) else None, default=0.0)
            note = _to_str(self.table.item(r, 3).text() if self.table.item(r, 3) else "")

            if note:
                notes.append(note)

            if not material:
                continue
            mats.append(MaterialRow(material=material, density_g_cm3=density, z_factor=zfac, note=note))

        # notes 중복 제거
        uniq = []
        seen = set()
        for n in notes:
            if n not in seen:
                uniq.append(n)
                seen.add(n)

        return mats, uniq

    def _save_json(self, mats: list[MaterialRow], notes: list[str]) -> None:
        obj = {
            "version": 1,
            "materials": [
                {
                    "material": m.material,
                    "density_g_cm3": float(m.density_g_cm3),
                    "z_factor": float(m.z_factor),
                    "note": m.note or "",
                }
                for m in mats
            ],
            "notes": notes,
        }
        self._json_path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

    def _validate_cell(self, col: int, text: str) -> bool:
        # 0: material (빈 값 금지)
        if col == 0:
            return bool(_to_str(text))

        # 1: density > 0
        if col == 1:
            v = _to_float(text, default=-1.0)
            return v > 0.0

        # 2: z_factor > 0
        if col == 2:
            v = _to_float(text, default=-1.0)
            return v > 0.0

        # 3: note는 자유
        return True

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if self._loading:
            return

        col = item.column()
        txt = item.text()

        if not self._validate_cell(col, txt):
            QMessageBox.warning(
                self,
                "Invalid value",
                "값이 올바르지 않습니다.\n"
                "- Material: 빈 값 불가\n"
                "- Density: 0보다 큰 숫자\n"
                "- Z factor: 0보다 큰 숫자",
            )
            # 유효하지 않으면 일단 빈칸으로 되돌려 강제 저장 방지
            self._loading = True
            item.setText("")
            self._loading = False
            return

        mats, notes = self._collect_rows()
        self._save_json(mats, notes)

    def _on_apply(self) -> None:
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Select", "적용할 행을 먼저 선택하세요.")
            return

        material = _to_str(self.table.item(row, 0).text() if self.table.item(row, 0) else "")
        density = _to_float(self.table.item(row, 1).text() if self.table.item(row, 1) else None, default=0.0)
        zfac = _to_float(self.table.item(row, 2).text() if self.table.item(row, 2) else None, default=0.0)
        note = _to_str(self.table.item(row, 3).text() if self.table.item(row, 3) else "")

        if not material:
            QMessageBox.warning(self, "Invalid", "Material이 비어있습니다.")
            return
        if density <= 0 or zfac <= 0:
            QMessageBox.warning(self, "Invalid", "Density/Z factor 값이 올바르지 않습니다.")
            return

        # 최신 테이블 상태 저장
        mats, notes = self._collect_rows()
        self._save_json(mats, notes)

        self._selected = MaterialRow(material=material, density_g_cm3=density, z_factor=zfac, note=note)
        self.accept()
